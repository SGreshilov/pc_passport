from openpyxl import load_workbook


wb = load_workbook(filename='source\\asd.xlsx', read_only=True)
# wb = load_workbook(filename='\\\\Srv-nfs\\пз\\ДИТ\\УИТ\\ОТОИС\\инвентаризация\\Учет СВТ.xlsx', read_only=True)
list_one = wb['Лист1']

def find_items(find):
    res = []
    for row in list_one.iter_rows(min_row=1, min_col=1, max_col=8):
        if isinstance(row[4].value, str) and find == row[4].value.lower() and row[2].value in ['монитор', 'системный блок']:
            res.append(tuple(str(cell.value) for cell in row))
    for row in res:
        for item in row:
            print(f'({item})', end=' ')
        print()
    check_items(res)
    return res


def check_items(table):
    if len(table) == 0:
        raise ValueError('В списке нет элементов, возможно вы неправильно указали имя компьютера')
    if len(table) == 1:
        raise ValueError('В списке находится один элемент')
    if len(table) > 4:
        raise ValueError('В списке находится больше четырех элементов')
    pc_cnt = 0
    for row in table:
        if row[2] == 'системный блок':
            pc_cnt += 1
    if pc_cnt > 1:
        raise ValueError('В списке находится больше одного ситемного блока')
    if pc_cnt == 0:
        raise ValueError('В списке отсутсвуют системные блоки')
    print('Проверка количества элементов прошла успешно')
